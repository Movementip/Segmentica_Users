import io
import base64
import hashlib
import json
import os
import posixpath
import shutil
import subprocess
import tempfile
import traceback
import unicodedata
import xml.etree.ElementTree as ET
import zipfile
from copy import copy, deepcopy
from pathlib import Path
from urllib.parse import quote
from flask import Flask, jsonify, request, send_file
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import (
    column_index_from_string,
    coordinate_to_tuple,
    get_column_letter,
    range_boundaries,
)
from openpyxl.worksheet.pagebreak import Break, RowBreak
from openpyxl.worksheet.properties import PageSetupProperties

app = Flask(__name__)

TEMPLATES_DIR = Path(os.environ.get("TEMPLATES_DIR", "/app/templates/forms")).resolve()
LIBREOFFICE_BIN = os.environ.get("LIBREOFFICE_BIN", "/usr/bin/soffice")
LIBREOFFICE_TIMEOUT_MS = int(os.environ.get("LIBREOFFICE_TIMEOUT_MS", "30000"))
CACHE_DIR = Path(os.environ.get("DOCUMENT_RENDERER_CACHE_DIR", "/tmp/segmentica-render-cache")).resolve()
SUPPORTED_POSTPROCESS = {"none", "stack_pages_vertical"}
WORDPROCESSING_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
DRAWING_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
WORD_DRAWING_NS = "http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing"
OFFICE_DOCUMENT_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
XML_NS = "http://www.w3.org/XML/1998/namespace"
WORD_PARAGRAPH_TAG = f"{{{WORDPROCESSING_NS}}}p"
WORD_PARAGRAPH_PROPERTIES_TAG = f"{{{WORDPROCESSING_NS}}}pPr"
WORD_RUN_TAG = f"{{{WORDPROCESSING_NS}}}r"
WORD_RUN_PROPERTIES_TAG = f"{{{WORDPROCESSING_NS}}}rPr"
WORD_JUSTIFICATION_TAG = f"{{{WORDPROCESSING_NS}}}jc"
WORD_TEXT_TAG = f"{{{WORDPROCESSING_NS}}}t"
WORD_TAB_TAG = f"{{{WORDPROCESSING_NS}}}tab"
WORD_BREAK_TAG = f"{{{WORDPROCESSING_NS}}}br"
DRAWING_BLIP_TAG = f"{{{DRAWING_NS}}}blip"
WORD_DRAWING_INLINE_TAG = f"{{{WORD_DRAWING_NS}}}inline"
WORD_DRAWING_ANCHOR_TAG = f"{{{WORD_DRAWING_NS}}}anchor"
RELATIONSHIP_TAG = f"{{{PACKAGE_REL_NS}}}Relationship"
REL_EMBED_ATTR = f"{{{OFFICE_DOCUMENT_REL_NS}}}embed"
FIRST_REPLACED_IMAGE_SCALE = 1.0
DOCX_MIMETYPE = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"

ET.register_namespace("w", WORDPROCESSING_NS)


def to_ascii_filename(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    cleaned = "".join(ch if ch.isalnum() or ch in {" ", ".", "-", "_", "(", ")"} else "_" for ch in ascii_only)
    cleaned = " ".join(cleaned.split()).strip(" .")
    return cleaned or "document"


def apply_download_filename(response, filename: str, as_attachment: bool = True):
    disposition = "attachment" if as_attachment else "inline"
    safe_filename = str(filename or "document").replace("\\", "_").replace('"', "'").replace("\r", " ").replace("\n", " ").strip()
    ascii_fallback = to_ascii_filename(safe_filename)
    encoded = quote(safe_filename, safe="")
    response.headers["Content-Disposition"] = (
        f"{disposition}; filename=\"{ascii_fallback}\"; filename*=UTF-8''{encoded}"
    )
    return response


def ensure_template_path(template_name: str) -> Path:
    target = (TEMPLATES_DIR / template_name).resolve()
    if not str(target).startswith(str(TEMPLATES_DIR)):
        raise ValueError("Template path is outside templates directory")
    if not target.exists():
        raise FileNotFoundError(f"Template not found: {template_name}")
    return target


def load_workbook_with_fallback(template_path: Path):
    first_error = None
    try:
        workbook = load_workbook(template_path)
        if workbook.sheetnames:
            return workbook
        first_error = RuntimeError("openpyxl loaded workbook without visible sheets")
    except Exception as error:  # noqa: BLE001
        first_error = error

    temp_dir = Path(tempfile.mkdtemp(prefix="segmentica-template-normalize-"))
    try:
        user_installation_dir = temp_dir / "lo-profile"
        user_installation_dir.mkdir(parents=True, exist_ok=True)
        proc = subprocess.run(
            [
                LIBREOFFICE_BIN,
                "--headless",
                "--nologo",
                "--nolockcheck",
                "--nodefault",
                "--norestore",
                f"-env:UserInstallation=file://{user_installation_dir}",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(temp_dir),
                str(template_path),
            ],
            capture_output=True,
            text=True,
            timeout=LIBREOFFICE_TIMEOUT_MS / 1000,
            check=False,
        )

        if proc.returncode != 0:
            raise RuntimeError(
                f"LibreOffice normalize failed with code {proc.returncode}: {proc.stderr.strip() or proc.stdout.strip() or 'no output'}"
            ) from first_error

        normalized_path = temp_dir / f"{template_path.stem}.xlsx"
        if not normalized_path.exists():
            normalized_candidates = sorted(
                [
                    path
                    for path in temp_dir.glob("*.xlsx")
                    if path.is_file()
                ],
                key=lambda path: path.stat().st_mtime,
                reverse=True,
            )
            normalized_path = normalized_candidates[0] if normalized_candidates else None

        if normalized_path is None or not normalized_path.exists():
            raise RuntimeError("LibreOffice did not produce a normalized XLSX file") from first_error

        workbook = load_workbook(normalized_path)
        if not workbook.sheetnames:
            raise RuntimeError("Normalized workbook still has no sheets") from first_error
        return workbook
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def copy_range_between_sheets(
    workbook,
    source_sheet_name: str,
    source_range: str,
    target_sheet_name: str,
    target_start_address: str,
) -> None:
    source_ws = workbook[source_sheet_name]
    target_ws = workbook[target_sheet_name]

    min_col, min_row, max_col, max_row = range_boundaries(source_range)
    target_col = column_index_from_string("".join(ch for ch in target_start_address if ch.isalpha()))
    target_row = int("".join(ch for ch in target_start_address if ch.isdigit()))
    row_offset = target_row - min_row
    col_offset = target_col - min_col

    for row in range(min_row, max_row + 1):
        if source_ws.row_dimensions[row].height is not None:
            target_ws.row_dimensions[row + row_offset].height = source_ws.row_dimensions[row].height
        target_ws.row_dimensions[row + row_offset].hidden = source_ws.row_dimensions[row].hidden

    target_min_row = target_row
    target_max_row = target_row + (max_row - min_row)
    target_min_col = target_col
    target_max_col = target_col + (max_col - min_col)

    for merged in list(target_ws.merged_cells.ranges):
        merged_min_col, merged_min_row, merged_max_col, merged_max_row = range_boundaries(str(merged))
        if (
            merged_min_row <= target_max_row
            and merged_max_row >= target_min_row
            and merged_min_col <= target_max_col
            and merged_max_col >= target_min_col
        ):
            target_ws.unmerge_cells(str(merged))

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            source_cell = source_ws.cell(row=row, column=col)
            if isinstance(source_cell, MergedCell):
                continue
            target_address = f"{get_column_letter(col + col_offset)}{row + row_offset}"
            target_cell = set_cell_value(target_ws, target_address, source_cell.value)
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            target_cell._style = copy(source_cell._style)

    for merged in source_ws.merged_cells.ranges:
        merged_min_col, merged_min_row, merged_max_col, merged_max_row = range_boundaries(str(merged))
        if (
            merged_min_row >= min_row
            and merged_max_row <= max_row
            and merged_min_col >= min_col
            and merged_max_col <= max_col
        ):
            target_ws.merge_cells(
                start_row=merged_min_row + row_offset,
                start_column=merged_min_col + col_offset,
                end_row=merged_max_row + row_offset,
                end_column=merged_max_col + col_offset,
            )


def copy_sheet(workbook, source_sheet_name: str, target_sheet_name: str) -> None:
    if source_sheet_name not in workbook.sheetnames:
        return
    if target_sheet_name in workbook.sheetnames:
        return

    source_ws = workbook[source_sheet_name]
    cloned_ws = workbook.copy_worksheet(source_ws)
    cloned_ws.title = target_sheet_name


def get_writable_cell(worksheet, address: str):
    cell = worksheet[address]
    if not isinstance(cell, MergedCell):
        return cell

    row, col = coordinate_to_tuple(address)
    for merged in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged))
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return worksheet.cell(row=min_row, column=min_col)

    return worksheet.cell(row=row, column=col)


def set_cell_value(worksheet, address: str, value):
    try:
        cell = get_writable_cell(worksheet, address)
        cell.value = value
        return cell
    except AttributeError as error:
        if "MergedCell" not in str(error):
            raise

        row, col = coordinate_to_tuple(address)
        for merged in list(worksheet.merged_cells.ranges):
            min_col, min_row, max_col, max_row = range_boundaries(str(merged))
            if min_row <= row <= max_row and min_col <= col <= max_col:
                top_left = worksheet.cell(row=min_row, column=min_col)
                top_left.value = value
                return top_left

        raise RuntimeError(f"Failed to write merged cell value at {worksheet.title}!{address}") from error


def apply_cells_to_workbook(
    template_path: Path,
    output_path: Path,
    cells: list[dict],
    row_visibility: list[dict],
    row_heights: list[dict],
    row_breaks: list[dict],
    print_areas: list[dict],
    range_copies: list[dict],
    sheet_copies: list[dict],
    hidden_sheets: list[str],
    sheet_page_setup: list[dict],
) -> None:
    workbook = load_workbook_with_fallback(template_path)

    for item in sheet_copies:
        source_sheet_name = str(item.get("sourceSheetName") or "").strip()
        target_sheet_name = str(item.get("targetSheetName") or "").strip()
        if not source_sheet_name or not target_sheet_name:
            continue
        copy_sheet(workbook, source_sheet_name, target_sheet_name)

    for item in range_copies:
        source_sheet_name = str(item.get("sourceSheetName") or "").strip()
        source_range = str(item.get("sourceRange") or "").strip()
        target_sheet_name = str(item.get("targetSheetName") or "").strip()
        target_start_address = str(item.get("targetStartAddress") or "").strip()
        if not source_sheet_name or not source_range or not target_sheet_name or not target_start_address:
            continue
        if source_sheet_name not in workbook.sheetnames or target_sheet_name not in workbook.sheetnames:
            continue
        copy_range_between_sheets(
            workbook,
            source_sheet_name,
            source_range,
            target_sheet_name,
            target_start_address,
        )

    for item in cells:
        address = str(item.get("address") or "").strip()
        if not address:
            continue
        sheet_name = str(item.get("sheetName") or "").strip()
        worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        cell = set_cell_value(worksheet, address, item.get("value"))

        style = item.get("style") or {}
        if isinstance(style, dict) and style:
            font = copy(cell.font)
            alignment = copy(cell.alignment)

            if style.get("fontName"):
                font.name = str(style.get("fontName"))
            if style.get("fontSize") is not None:
                font.sz = float(style.get("fontSize"))
            if style.get("bold") is not None:
                font.bold = bool(style.get("bold"))

            if style.get("horizontal"):
                alignment.horizontal = str(style.get("horizontal"))
            if style.get("vertical"):
                alignment.vertical = str(style.get("vertical"))
            if style.get("wrapText") is not None:
                alignment.wrap_text = bool(style.get("wrapText"))
            if style.get("shrinkToFit") is not None:
                alignment.shrink_to_fit = bool(style.get("shrinkToFit"))

            cell.font = font
            cell.alignment = alignment

    for item in row_visibility:
        sheet_name = str(item.get("sheetName") or "").strip()
        row = int(item.get("row") or 0)
        hidden = bool(item.get("hidden"))
        if row <= 0:
            continue
        worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        worksheet.row_dimensions[row].hidden = hidden

    for item in row_heights:
        sheet_name = str(item.get("sheetName") or "").strip()
        row = int(item.get("row") or 0)
        height = item.get("height")
        if row <= 0 or height is None:
            continue
        worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        worksheet.row_dimensions[row].height = float(height)

    for item in row_breaks:
        sheet_name = str(item.get("sheetName") or "").strip()
        worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        if bool(item.get("clearExisting")):
            worksheet.row_breaks = RowBreak()

        breaks = item.get("breaks") or []
        if not isinstance(breaks, list):
            continue

        for row in breaks:
            try:
                row_number = int(row)
            except (TypeError, ValueError):
                continue
            if row_number <= 0:
                continue
            worksheet.row_breaks.append(Break(id=row_number))

    for item in print_areas:
        sheet_name = str(item.get("sheetName") or "").strip()
        area = str(item.get("range") or "").strip()
        if not area:
            continue
        worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        worksheet.print_area = area

    for sheet_name in list(hidden_sheets):
        if sheet_name in workbook.sheetnames and len(workbook.sheetnames) > 1:
            worksheet = workbook[sheet_name]
            workbook.remove(worksheet)

    page_setup_map = {}
    for item in sheet_page_setup:
        sheet_name = str(item.get("sheetName") or "").strip()
        if not sheet_name:
            continue
        page_setup_map[sheet_name] = {
            "fitToWidth": item.get("fitToWidth"),
            "fitToHeight": item.get("fitToHeight"),
        }

    for sheet in workbook.worksheets:
        # We want each Excel sheet to become one PDF page before stacking.
        # Otherwise LibreOffice preserves internal page breaks and we end up
        # stitching many fragmented pages instead of whole worksheets.
        sheet.sheet_view.view = "normal"
        if sheet.sheet_properties.pageSetUpPr is None:
            sheet.sheet_properties.pageSetUpPr = PageSetupProperties()
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        page_setup = page_setup_map.get(sheet.title, {})
        fit_to_width = page_setup.get("fitToWidth")
        fit_to_height = page_setup.get("fitToHeight")
        sheet.page_setup.fitToWidth = 1 if fit_to_width is None else int(fit_to_width)
        sheet.page_setup.fitToHeight = 1 if fit_to_height is None else int(fit_to_height)
        if not sheet.print_area:
            sheet.print_area = sheet.calculate_dimension()

    workbook.save(output_path)


def should_normalize_excel_output(template_name: str, output_format: str) -> bool:
    normalized_name = str(template_name or "").lower()
    normalized_format = str(output_format or "").lower()
    return normalized_format == "excel" and "т-13" in normalized_name and "табель" in normalized_name


def normalize_excel_output(workbook_path: Path, temp_dir: Path) -> Path:
    normalize_dir = temp_dir / "xlsx-normalized"
    normalize_dir.mkdir(parents=True, exist_ok=True)
    user_installation_dir = normalize_dir / "lo-profile"
    user_installation_dir.mkdir(parents=True, exist_ok=True)

    proc = subprocess.run(
        [
            LIBREOFFICE_BIN,
            "--headless",
            "--nologo",
            "--nolockcheck",
            "--nodefault",
            "--norestore",
            f"-env:UserInstallation=file://{user_installation_dir}",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(normalize_dir),
            str(workbook_path),
        ],
        capture_output=True,
        text=True,
        timeout=LIBREOFFICE_TIMEOUT_MS / 1000,
        check=False,
    )

    if proc.returncode != 0:
        raise RuntimeError(
            f"LibreOffice Excel normalize failed with code {proc.returncode}: {proc.stderr.strip() or proc.stdout.strip() or 'no output'}"
        )

    normalized_path = normalize_dir / workbook_path.name
    if not normalized_path.exists():
        raise RuntimeError("LibreOffice did not produce a normalized Excel file")

    return normalized_path


def convert_to_pdf(input_path: Path, output_dir: Path) -> Path:
    user_installation_dir = output_dir / "lo-profile"
    user_installation_dir.mkdir(parents=True, exist_ok=True)

    proc = subprocess.run(
        [
            LIBREOFFICE_BIN,
            "--headless",
            "--nologo",
            "--nolockcheck",
            "--nodefault",
            "--norestore",
            f"-env:UserInstallation=file://{user_installation_dir}",
            "--convert-to",
            "pdf",
            "--outdir",
            str(output_dir),
            str(input_path),
        ],
        capture_output=True,
        text=True,
        timeout=LIBREOFFICE_TIMEOUT_MS / 1000,
        check=False,
    )

    if proc.returncode != 0:
        raise RuntimeError(
            f"LibreOffice convert failed with code {proc.returncode}: {proc.stderr.strip() or proc.stdout.strip() or 'no output'}"
        )

    pdf_path = output_dir / f"{input_path.stem}.pdf"
    if not pdf_path.exists():
        raise RuntimeError("LibreOffice did not produce a PDF file")
    return pdf_path


def convert_word_template_to_docx(input_path: Path, output_dir: Path) -> Path:
    if input_path.suffix.lower() == ".docx":
        return input_path

    user_installation_dir = output_dir / "lo-profile-word-normalize"
    user_installation_dir.mkdir(parents=True, exist_ok=True)

    proc = subprocess.run(
        [
            LIBREOFFICE_BIN,
            "--headless",
            "--nologo",
            "--nolockcheck",
            "--nodefault",
            "--norestore",
            f"-env:UserInstallation=file://{user_installation_dir}",
            "--convert-to",
            "docx",
            "--outdir",
            str(output_dir),
            str(input_path),
        ],
        capture_output=True,
        text=True,
        timeout=LIBREOFFICE_TIMEOUT_MS / 1000,
        check=False,
    )

    if proc.returncode != 0:
        raise RuntimeError(
            f"LibreOffice Word normalize failed with code {proc.returncode}: {proc.stderr.strip() or proc.stdout.strip() or 'no output'}"
        )

    normalized_path = output_dir / f"{input_path.stem}.docx"
    if not normalized_path.exists():
        candidates = sorted(
            [path for path in output_dir.glob("*.docx") if path.is_file()],
            key=lambda path: path.stat().st_mtime,
            reverse=True,
        )
        normalized_path = candidates[0] if candidates else None

    if normalized_path is None or not normalized_path.exists():
        raise RuntimeError("LibreOffice did not produce a normalized DOCX file")

    return normalized_path


def _build_parent_map(root):
    return {child: parent for parent in root.iter() for child in parent}


def _append_text_node(run, value: str) -> None:
    text_node = ET.Element(WORD_TEXT_TAG)
    if value.startswith(" ") or value.endswith(" ") or "  " in value:
        text_node.set(f"{{{XML_NS}}}space", "preserve")
    text_node.text = value
    run.append(text_node)


def _set_run_text(run, text: str) -> None:
    run_properties = run.find(WORD_RUN_PROPERTIES_TAG)
    for child in list(run):
        if child is not run_properties:
            run.remove(child)

    lines = str(text or "").split("\n")
    if not lines:
        lines = [""]
    for index, line in enumerate(lines):
        segments = str(line).split("\t")
        for segment_index, segment in enumerate(segments):
            _append_text_node(run, segment)
            if segment_index < len(segments) - 1:
                run.append(ET.Element(WORD_TAB_TAG))
        if index < len(lines) - 1:
            run.append(ET.Element(WORD_BREAK_TAG))


def _set_paragraph_text(paragraph, text: str) -> None:
    runs = [run for run in paragraph.iter(WORD_RUN_TAG)]
    if not runs:
        run = ET.Element(WORD_RUN_TAG)
        paragraph.append(run)
        runs = [run]

    parent_map = _build_parent_map(paragraph)
    first_run = runs[0]
    _set_run_text(first_run, text)

    for run in runs[1:]:
        parent = parent_map.get(run)
        if parent is not None:
            parent.remove(run)


def _set_paragraph_alignment(paragraph, alignment: str | None) -> None:
    paragraph_properties = paragraph.find(WORD_PARAGRAPH_PROPERTIES_TAG)
    if paragraph_properties is None:
        paragraph_properties = ET.Element(WORD_PARAGRAPH_PROPERTIES_TAG)
        paragraph.insert(0, paragraph_properties)

    justification = paragraph_properties.find(WORD_JUSTIFICATION_TAG)
    if alignment:
        if justification is None:
            justification = ET.Element(WORD_JUSTIFICATION_TAG)
            paragraph_properties.append(justification)
        justification.set(f"{{{WORDPROCESSING_NS}}}val", alignment)
    elif justification is not None:
        paragraph_properties.remove(justification)


def _force_paragraph_alignment(paragraph, alignment: str | None) -> None:
    paragraph_properties = paragraph.find(WORD_PARAGRAPH_PROPERTIES_TAG)
    if paragraph_properties is not None:
        paragraph.remove(paragraph_properties)
    paragraph_properties = ET.Element(WORD_PARAGRAPH_PROPERTIES_TAG)
    paragraph.insert(0, paragraph_properties)
    if alignment:
        justification = ET.Element(WORD_JUSTIFICATION_TAG)
        justification.set(f"{{{WORDPROCESSING_NS}}}val", alignment)
        paragraph_properties.append(justification)


def _copy_paragraph_properties(source_paragraph, target_paragraph) -> None:
    source_properties = source_paragraph.find(WORD_PARAGRAPH_PROPERTIES_TAG)
    target_properties = target_paragraph.find(WORD_PARAGRAPH_PROPERTIES_TAG)
    if target_properties is not None:
        target_paragraph.remove(target_properties)
    if source_properties is not None:
        target_paragraph.insert(0, deepcopy(source_properties))


def _replace_paragraph_with_source(source_paragraph, target_paragraph, text: str, alignment: str = None) -> None:
    _copy_paragraph_properties(source_paragraph, target_paragraph)
    _set_paragraph_text(target_paragraph, text)
    if alignment is not None:
        _force_paragraph_alignment(target_paragraph, alignment)


def _run_contains_drawing(run) -> bool:
    return any(child.tag == f"{{{WORDPROCESSING_NS}}}drawing" for child in list(run))


def _set_paragraph_text_preserving_drawings(paragraph, text: str) -> None:
    runs = [child for child in list(paragraph) if child.tag == WORD_RUN_TAG]
    if not runs:
        _set_paragraph_text(paragraph, text)
        return

    target_run = next((run for run in runs if not _run_contains_drawing(run)), None)
    if target_run is None:
        target_run = ET.Element(WORD_RUN_TAG)
        paragraph_properties = paragraph.find(WORD_PARAGRAPH_PROPERTIES_TAG)
        insert_index = 1 if paragraph_properties is not None else 0
        paragraph.insert(insert_index, target_run)
        runs = [child for child in list(paragraph) if child.tag == WORD_RUN_TAG]

    removable_runs = []
    for run in runs:
        if _run_contains_drawing(run):
            continue
        if run is target_run:
            continue
        removable_runs.append(run)

    _set_run_text(target_run, text)
    for run in removable_runs:
        paragraph.remove(run)


def _remove_drawing_runs(paragraph) -> None:
    for run in [child for child in list(paragraph) if child.tag == WORD_RUN_TAG]:
        if _run_contains_drawing(run):
            paragraph.remove(run)


def _set_table_cell_text(cell, text: str, paragraph_index: int = 0) -> None:
    paragraphs = [child for child in list(cell) if child.tag == WORD_PARAGRAPH_TAG]
    if not paragraphs:
        paragraph = ET.Element(WORD_PARAGRAPH_TAG)
        cell.append(paragraph)
        paragraphs = [paragraph]

    while len(paragraphs) <= paragraph_index:
        paragraph = ET.Element(WORD_PARAGRAPH_TAG)
        cell.append(paragraph)
        paragraphs.append(paragraph)

    _set_paragraph_text(paragraphs[paragraph_index], text)

    for extra in paragraphs[paragraph_index + 1:]:
        _set_paragraph_text(extra, "")


def _set_table_cell_alignment(cell, alignment: str | None, paragraph_index: int = 0) -> None:
    paragraphs = [child for child in list(cell) if child.tag == WORD_PARAGRAPH_TAG]
    if not paragraphs:
        paragraph = ET.Element(WORD_PARAGRAPH_TAG)
        cell.append(paragraph)
        paragraphs = [paragraph]

    while len(paragraphs) <= paragraph_index:
        paragraph = ET.Element(WORD_PARAGRAPH_TAG)
        cell.append(paragraph)
        paragraphs.append(paragraph)

    _set_paragraph_alignment(paragraphs[paragraph_index], alignment)


def _force_table_cell_alignment(cell, alignment: str | None, paragraph_index: int = 0) -> None:
    paragraphs = [child for child in list(cell) if child.tag == WORD_PARAGRAPH_TAG]
    if not paragraphs:
        paragraph = ET.Element(WORD_PARAGRAPH_TAG)
        cell.append(paragraph)
        paragraphs = [paragraph]

    while len(paragraphs) <= paragraph_index:
        paragraph = ET.Element(WORD_PARAGRAPH_TAG)
        cell.append(paragraph)
        paragraphs.append(paragraph)

    _force_paragraph_alignment(paragraphs[paragraph_index], alignment)


def _copy_table_cell_paragraph_properties(source_cell, target_cell, source_paragraph_index: int = 0, target_paragraph_index: int = 0) -> None:
    source_paragraphs = [child for child in list(source_cell) if child.tag == WORD_PARAGRAPH_TAG]
    target_paragraphs = [child for child in list(target_cell) if child.tag == WORD_PARAGRAPH_TAG]
    if not source_paragraphs:
        return
    while len(target_paragraphs) <= target_paragraph_index:
        paragraph = ET.Element(WORD_PARAGRAPH_TAG)
        target_cell.append(paragraph)
        target_paragraphs.append(paragraph)
    _copy_paragraph_properties(source_paragraphs[source_paragraph_index], target_paragraphs[target_paragraph_index])


def apply_order_supply_specification_template(
    template_path: Path,
    output_path: Path,
    replacements: dict[str, str],
) -> None:
    rows_json = str(replacements.get("__SPECIFICATION_ROWS_JSON__") or "[]")
    try:
        rows = json.loads(rows_json)
    except json.JSONDecodeError as error:
        raise RuntimeError(f"Invalid specification rows payload: {error}") from error

    with zipfile.ZipFile(template_path, "r") as template_zip, zipfile.ZipFile(
        output_path, "w", compression=zipfile.ZIP_DEFLATED
    ) as output_zip:
        for entry in template_zip.infolist():
            data = template_zip.read(entry.filename)
            if entry.filename == "word/document.xml":
                root = ET.fromstring(data)
                body = root.find(f"{{{WORDPROCESSING_NS}}}body")
                if body is None:
                    raise RuntimeError("Specification template has no document body")

                paragraphs = [
                    child for child in list(body)
                    if child.tag == WORD_PARAGRAPH_TAG
                ]
                non_empty_paragraphs = []
                for paragraph in paragraphs:
                    text = "".join(
                        text_node.text or ""
                        for text_node in paragraph.iter(WORD_TEXT_TAG)
                    ).strip()
                    if text:
                        non_empty_paragraphs.append(paragraph)

                if len(non_empty_paragraphs) >= 5:
                    _set_paragraph_text(non_empty_paragraphs[2], str(replacements.get("__SPECIFICATION_HEADER_BASIS__") or ""))
                    _force_paragraph_alignment(non_empty_paragraphs[2], "right")
                    _set_paragraph_text(non_empty_paragraphs[3], str(replacements.get("__SPECIFICATION_TITLE__") or ""))
                    _force_paragraph_alignment(non_empty_paragraphs[3], "center")
                    _set_paragraph_text(non_empty_paragraphs[4], str(replacements.get("__SPECIFICATION_DATE_LONG__") or ""))
                    _force_paragraph_alignment(non_empty_paragraphs[4], "right")

                tables = root.findall(f".//{{{WORDPROCESSING_NS}}}tbl")
                if len(tables) < 2:
                    raise RuntimeError("Specification template structure was not recognized")

                items_table = tables[0]
                item_rows = items_table.findall(f"./{{{WORDPROCESSING_NS}}}tr")
                if len(item_rows) < 3:
                    raise RuntimeError("Specification items table is too short")

                header_row = item_rows[0]
                sample_row = item_rows[1]
                total_row = item_rows[-1]

                for row in item_rows[1:-1]:
                    items_table.remove(row)

                insert_at = list(items_table).index(total_row)
                normalized_rows = rows if isinstance(rows, list) and rows else []
                if not normalized_rows:
                    normalized_rows = [{
                        "number": "1",
                        "name": "—",
                        "quantity": "0",
                        "price": "0,00",
                    }]

                for row_data in normalized_rows:
                    new_row = deepcopy(sample_row)
                    cells = new_row.findall(f"./{{{WORDPROCESSING_NS}}}tc")
                    if len(cells) >= 4:
                        _set_table_cell_text(cells[0], str(row_data.get("number") or ""))
                        _set_table_cell_text(cells[1], str(row_data.get("name") or ""))
                        _set_table_cell_text(cells[2], str(row_data.get("quantity") or ""))
                        _set_table_cell_text(cells[3], str(row_data.get("price") or ""))
                    items_table.insert(insert_at, new_row)
                    insert_at += 1

                total_cells = total_row.findall(f"./{{{WORDPROCESSING_NS}}}tc")
                if len(total_cells) >= 2:
                    _set_table_cell_text(total_cells[0], "Итого:")
                    _force_table_cell_alignment(total_cells[0], "right")
                    _force_table_cell_alignment(total_cells[1], "left")
                    sample_cells = sample_row.findall(f"./{{{WORDPROCESSING_NS}}}tc")
                    if len(sample_cells) >= 3:
                        source_paragraphs = sample_cells[2].findall(f"./{{{WORDPROCESSING_NS}}}p")
                        target_paragraphs = total_cells[1].findall(f"./{{{WORDPROCESSING_NS}}}p")
                        if source_paragraphs and target_paragraphs:
                            _replace_paragraph_with_source(
                                source_paragraphs[0],
                                target_paragraphs[0],
                                str(replacements.get("__SPECIFICATION_TOTAL__") or ""),
                                "left",
                            )
                        else:
                            _set_table_cell_text(total_cells[1], str(replacements.get("__SPECIFICATION_TOTAL__") or ""))
                            _force_table_cell_alignment(total_cells[1], "left")
                    else:
                        _set_table_cell_text(total_cells[1], str(replacements.get("__SPECIFICATION_TOTAL__") or ""))
                        _force_table_cell_alignment(total_cells[1], "left")

                parties_table = tables[1]
                parties_rows = parties_table.findall(f"./{{{WORDPROCESSING_NS}}}tr")
                if parties_rows:
                    parties_cells = parties_rows[0].findall(f"./{{{WORDPROCESSING_NS}}}tc")
                    if len(parties_cells) >= 2:
                        _set_table_cell_text(parties_cells[0], "Поставщик:", 0)
                        supplier_name = str(replacements.get("__SPECIFICATION_SUPPLIER_NAME__") or "")
                        supplier_position = str(replacements.get("__SPECIFICATION_SUPPLIER_POSITION__") or "")
                        supplier_fio = str(replacements.get("__SPECIFICATION_SUPPLIER_FIO__") or "")
                        supplier_line = supplier_position
                        if supplier_line:
                            supplier_line += " "
                        supplier_line += "__________"
                        if supplier_fio:
                            supplier_line += f" {supplier_fio}"
                        _set_table_cell_text(parties_cells[0], f"{supplier_name}\n{supplier_line}", 1)
                        _set_table_cell_text(parties_cells[0], "М.П.", 2)
                        _force_table_cell_alignment(parties_cells[0], None, 0)
                        _force_table_cell_alignment(parties_cells[0], None, 1)
                        _force_table_cell_alignment(parties_cells[0], "center", 2)

                        _set_table_cell_text(parties_cells[1], str(replacements.get("__SPECIFICATION_BUYER_LABEL__") or "Покупатель:"), 0)
                        buyer_name = str(replacements.get("__SPECIFICATION_BUYER_NAME__") or "")
                        buyer_position = str(replacements.get("__SPECIFICATION_BUYER_POSITION__") or "")
                        buyer_fio = str(replacements.get("__SPECIFICATION_BUYER_FIO__") or "")
                        buyer_line = buyer_position
                        if buyer_line:
                            buyer_line += " "
                        buyer_line += "__________"
                        if buyer_fio:
                            buyer_line += f" {buyer_fio}"
                        _set_table_cell_text(parties_cells[1], f"{buyer_name}\n{buyer_line}", 1)
                        _set_table_cell_text(parties_cells[1], "М.П.", 2)
                        _force_table_cell_alignment(parties_cells[1], None, 0)
                        _force_table_cell_alignment(parties_cells[1], None, 1)
                        _force_table_cell_alignment(parties_cells[1], "center", 2)

                data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            output_zip.writestr(entry, data)


def apply_structured_invoice_template(
    template_path: Path,
    output_path: Path,
    replacements: dict[str, str],
    replace_first_image_bytes: bytes = None,
) -> None:
    rows_json = str(replacements.get("__ALT_INVOICE_ROWS_JSON__") or "[]")
    try:
        rows = json.loads(rows_json)
    except json.JSONDecodeError as error:
        raise RuntimeError(f"Invalid structured invoice rows payload: {error}") from error

    with zipfile.ZipFile(template_path, "r") as template_zip, zipfile.ZipFile(
        output_path, "w", compression=zipfile.ZIP_DEFLATED
    ) as output_zip:
        first_image_entry_name = _resolve_first_image_entry_name(template_zip)

        for entry in template_zip.infolist():
            data = template_zip.read(entry.filename)
            if replace_first_image_bytes and entry.filename == first_image_entry_name:
                data = replace_first_image_bytes
            elif entry.filename == "word/document.xml":
                root = ET.fromstring(data)
                body = root.find(f"{{{WORDPROCESSING_NS}}}body")
                if body is None:
                    raise RuntimeError("Invoice template has no document body")

                body_children = list(body)
                if len(body_children) < 15:
                    raise RuntimeError("Invoice template structure was not recognized")

                company_name = str(replacements.get("__ALT_INVOICE_COMPANY_NAME__") or "")
                company_address = str(replacements.get("__ALT_INVOICE_COMPANY_ADDRESS__") or "")
                company_contacts = str(replacements.get("__ALT_INVOICE_COMPANY_CONTACTS__") or "")

                first_paragraph = body_children[0]
                if replace_first_image_bytes:
                    _set_paragraph_text_preserving_drawings(first_paragraph, company_name)
                else:
                    _remove_drawing_runs(first_paragraph)
                    _set_paragraph_text(first_paragraph, company_name)

                _set_paragraph_text(body_children[1], company_address)
                _set_paragraph_text(body_children[2], company_contacts)
                _set_paragraph_text(body_children[6], str(replacements.get("__ALT_INVOICE_TITLE__") or ""))
                _force_paragraph_alignment(body_children[6], "center")
                _set_paragraph_text(body_children[10], str(replacements.get("__ALT_INVOICE_TOTAL_LINE__") or ""))
                _force_paragraph_alignment(body_children[10], "right")
                _set_paragraph_text(body_children[11], str(replacements.get("__ALT_INVOICE_TOTAL_WORDS__") or ""))

                bank_table = body_children[4]
                bank_rows = bank_table.findall(f"./{{{WORDPROCESSING_NS}}}tr")
                if len(bank_rows) >= 6:
                    row0 = bank_rows[0].findall(f"./{{{WORDPROCESSING_NS}}}tc")
                    row1 = bank_rows[1].findall(f"./{{{WORDPROCESSING_NS}}}tc")
                    row3 = bank_rows[3].findall(f"./{{{WORDPROCESSING_NS}}}tc")
                    row4 = bank_rows[4].findall(f"./{{{WORDPROCESSING_NS}}}tc")
                    if len(row0) >= 3:
                        _set_table_cell_text(row0[0], str(replacements.get("__ALT_INVOICE_BANK_NAME__") or ""))
                        _set_table_cell_text(row0[2], str(replacements.get("__ALT_INVOICE_BIK__") or ""))
                    if len(row1) >= 3:
                        _set_table_cell_text(row1[2], str(replacements.get("__ALT_INVOICE_CORR_ACCOUNT__") or ""))
                    if len(row3) >= 6:
                        _set_table_cell_text(row3[1], str(replacements.get("__ALT_INVOICE_INN__") or ""))
                        _set_table_cell_text(row3[3], str(replacements.get("__ALT_INVOICE_KPP__") or ""))
                        _set_table_cell_text(row3[5], str(replacements.get("__ALT_INVOICE_SETTLEMENT_ACCOUNT__") or ""))
                    if len(row4) >= 1:
                        _set_table_cell_text(row4[0], str(replacements.get("__ALT_INVOICE_RECIPIENT__") or ""))

                counterparties_table = body_children[7]
                counterparty_rows = counterparties_table.findall(f"./{{{WORDPROCESSING_NS}}}tr")
                if len(counterparty_rows) >= 2:
                    supplier_cells = counterparty_rows[0].findall(f"./{{{WORDPROCESSING_NS}}}tc")
                    buyer_cells = counterparty_rows[1].findall(f"./{{{WORDPROCESSING_NS}}}tc")
                    if len(supplier_cells) >= 2:
                        _set_table_cell_text(supplier_cells[1], str(replacements.get("__ALT_INVOICE_SUPPLIER_NAME__") or ""))
                    if len(buyer_cells) >= 2:
                        _set_table_cell_text(buyer_cells[1], str(replacements.get("__ALT_INVOICE_BUYER_TEXT__") or ""))

                items_sdt = body_children[9]
                sdt_content = items_sdt.find(f"./{{{WORDPROCESSING_NS}}}sdtContent")
                if sdt_content is not None:
                    items_table = sdt_content.find(f"./{{{WORDPROCESSING_NS}}}tbl")
                    if items_table is not None:
                        item_rows = items_table.findall(f"./{{{WORDPROCESSING_NS}}}tr")
                        if len(item_rows) >= 2:
                            header_row = item_rows[0]
                            sample_row = item_rows[1]
                            for row in item_rows[1:]:
                                items_table.remove(row)

                            normalized_rows = rows if isinstance(rows, list) and rows else [{
                                "number": "1",
                                "name": "—",
                                "unit": "",
                                "quantity": "",
                                "price": "0,00",
                                "sum": "0,00",
                            }]

                            for row_data in normalized_rows:
                                new_row = deepcopy(sample_row)
                                cells = new_row.findall(f"./{{{WORDPROCESSING_NS}}}tc")
                                if len(cells) >= 6:
                                    _set_table_cell_text(cells[0], str(row_data.get("number") or ""))
                                    _set_table_cell_text(cells[1], str(row_data.get("name") or ""))
                                    _set_table_cell_text(cells[2], str(row_data.get("unit") or ""))
                                    _set_table_cell_text(cells[3], str(row_data.get("quantity") or ""))
                                    _set_table_cell_text(cells[4], str(row_data.get("price") or ""))
                                    _set_table_cell_text(cells[5], str(row_data.get("sum") or ""))
                                items_table.append(new_row)

                signature_table = body_children[14]
                signature_rows = signature_table.findall(f"./{{{WORDPROCESSING_NS}}}tr")
                if signature_rows:
                    signature_cells = signature_rows[0].findall(f"./{{{WORDPROCESSING_NS}}}tc")
                    if len(signature_cells) >= 5:
                        _set_table_cell_text(signature_cells[0], str(replacements.get("__ALT_INVOICE_SIGN_LABEL__") or "Поставщик"))
                        _set_table_cell_text(signature_cells[1], str(replacements.get("__ALT_INVOICE_SIGN_POSITION__") or ""))
                        _set_table_cell_text(signature_cells[4], str(replacements.get("__ALT_INVOICE_SIGN_FIO__") or ""))

                data = ET.tostring(root, encoding="utf-8", xml_declaration=True)

            output_zip.writestr(entry, data)


def _replace_placeholders_in_paragraph(paragraph, replacements: dict[str, str]) -> bool:
    runs = [run for run in paragraph.iter(WORD_RUN_TAG)]
    if not runs:
        return False

    full_text = "".join("".join(text_node.text or "" for text_node in run.iter(WORD_TEXT_TAG)) for run in runs)
    if not full_text:
        return False

    replaced_text = full_text
    for placeholder, value in replacements.items():
        if not placeholder:
            continue
        replaced_text = replaced_text.replace(str(placeholder), str(value or ""))

    if replaced_text == full_text:
        return False

    parent_map = _build_parent_map(paragraph)
    first_run = runs[0]
    _set_run_text(first_run, replaced_text)

    for run in runs[1:]:
        parent = parent_map.get(run)
        if parent is not None:
            parent.remove(run)

    return True


def _resolve_first_image_entry_name(template_zip):
    try:
        document_root = ET.fromstring(template_zip.read("word/document.xml"))
        relationships_root = ET.fromstring(template_zip.read("word/_rels/document.xml.rels"))
    except (KeyError, ET.ParseError):
        return None

    relationships = {}
    for rel in relationships_root.iter(RELATIONSHIP_TAG):
        rel_id = rel.get("Id")
        target = rel.get("Target")
        if rel_id and target:
            relationships[rel_id] = target

    for blip in document_root.iter(DRAWING_BLIP_TAG):
        rel_id = blip.get(REL_EMBED_ATTR)
        target = relationships.get(rel_id)
        if not target:
            continue
        entry_name = target.lstrip("/") if target.startswith("/") else posixpath.normpath(posixpath.join("word", target))
        if entry_name.startswith("word/media/"):
            return entry_name

    return None


def _scale_first_drawing_container(root, scale_factor: float) -> bool:
    parent_map = _build_parent_map(root)

    for blip in root.iter(DRAWING_BLIP_TAG):
        current = blip
        while current is not None:
            current = parent_map.get(current)
            if current is None:
                break
            if current.tag not in {WORD_DRAWING_INLINE_TAG, WORD_DRAWING_ANCHOR_TAG}:
                continue

            for node in current.iter():
                cx = node.get("cx")
                cy = node.get("cy")
                if cx and cx.isdigit():
                    node.set("cx", str(max(1, round(int(cx) * scale_factor))))
                if cy and cy.isdigit():
                    node.set("cy", str(max(1, round(int(cy) * scale_factor))))
            return True

    return False


def _replace_first_image(
    template_path: Path,
    output_path: Path,
    replacements: dict[str, str],
    replace_first_image_bytes: bytes,
) -> None:
    with zipfile.ZipFile(template_path, "r") as template_zip, zipfile.ZipFile(
        output_path, "w", compression=zipfile.ZIP_DEFLATED
    ) as output_zip:
        first_image_entry_name = _resolve_first_image_entry_name(template_zip)
        for entry in template_zip.infolist():
            data = template_zip.read(entry.filename)
            if entry.filename == first_image_entry_name:
                data = replace_first_image_bytes
            elif entry.filename.startswith("word/") and entry.filename.endswith(".xml"):
                try:
                    root = ET.fromstring(data)
                    changed = False
                    for paragraph in root.iter(WORD_PARAGRAPH_TAG):
                        changed = _replace_placeholders_in_paragraph(paragraph, replacements) or changed
                    if entry.filename == "word/document.xml":
                        changed = _scale_first_drawing_container(root, FIRST_REPLACED_IMAGE_SCALE) or changed
                    if changed:
                        data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                except ET.ParseError:
                    pass
            output_zip.writestr(entry, data)


def apply_replacements_to_docx(
    template_path: Path,
    output_path: Path,
    replacements: dict[str, str],
    replace_first_image_bytes: bytes = None,
) -> None:
    if replace_first_image_bytes:
        _replace_first_image(template_path, output_path, replacements, replace_first_image_bytes)
    else:
        with zipfile.ZipFile(template_path, "r") as template_zip, zipfile.ZipFile(
            output_path, "w", compression=zipfile.ZIP_DEFLATED
        ) as output_zip:
            for entry in template_zip.infolist():
                data = template_zip.read(entry.filename)
                if entry.filename.startswith("word/") and entry.filename.endswith(".xml"):
                    try:
                        root = ET.fromstring(data)
                        changed = False
                        for paragraph in root.iter(WORD_PARAGRAPH_TAG):
                            changed = _replace_placeholders_in_paragraph(paragraph, replacements) or changed
                        if changed:
                            data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                    except ET.ParseError:
                        pass

                output_zip.writestr(entry, data)


def ensure_cache_dir() -> None:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)


def clear_cache_dir() -> None:
    if not CACHE_DIR.exists():
        return

    for path in CACHE_DIR.iterdir():
        try:
            if path.is_dir():
                shutil.rmtree(path, ignore_errors=True)
            else:
                path.unlink(missing_ok=True)
        except Exception:  # noqa: BLE001
            continue


def build_cache_key(xlsx_bytes: bytes, output_format: str, postprocess: str) -> str:
    digest = hashlib.sha256()
    digest.update(xlsx_bytes)
    digest.update(b"\0")
    digest.update(output_format.encode("utf-8"))
    digest.update(b"\0")
    digest.update(postprocess.encode("utf-8"))
    return digest.hexdigest()


@app.get("/health")
def health():
    return jsonify({
        "ok": True,
        "templatesDir": str(TEMPLATES_DIR),
        "libreofficeBin": LIBREOFFICE_BIN,
    })


@app.post("/convert/office-to-pdf")
def convert_office_to_pdf():
    uploaded_file = request.files.get("file")
    if uploaded_file is None or not uploaded_file.filename:
        return jsonify({"error": "file is required"}), 400

    original_name = Path(uploaded_file.filename).name
    original_stem = Path(original_name).stem or "document"
    original_suffix = Path(original_name).suffix or ".docx"
    safe_input_name = f"{to_ascii_filename(original_stem)}{original_suffix}"

    temp_dir = Path(tempfile.mkdtemp(prefix="segmentica-office-convert-"))
    try:
        input_path = temp_dir / safe_input_name
        uploaded_file.save(input_path)
        pdf_path = convert_to_pdf(input_path, temp_dir)
        pdf_bytes = pdf_path.read_bytes()

        response = send_file(
            io.BytesIO(pdf_bytes),
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"{original_stem}.pdf",
        )
        return apply_download_filename(response, f"{original_stem}.pdf")
    except Exception as error:  # noqa: BLE001
        traceback.print_exc()
        return jsonify({"error": str(error)}), 500
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


@app.post("/render/xlsx-template")
def render_xlsx_template():
    payload = request.get_json(silent=True) or {}
    template_name = str(payload.get("templateName") or "").strip()
    file_base_name = str(payload.get("fileBaseName") or "document").strip() or "document"
    output_format = str(payload.get("outputFormat") or "pdf").strip().lower()
    postprocess = str(payload.get("postprocess") or "").strip().lower()
    if not postprocess:
        postprocess = "stack_pages_vertical" if bool(payload.get("stackPages")) else "none"
    cells = payload.get("cells") or []
    row_visibility = payload.get("rowVisibility") or []
    row_heights = payload.get("rowHeights") or []
    row_breaks = payload.get("rowBreaks") or []
    print_areas = payload.get("printAreas") or []
    range_copies = payload.get("rangeCopies") or []
    sheet_copies = payload.get("sheetCopies") or []
    hidden_sheets = payload.get("hiddenSheets") or []
    sheet_page_setup = payload.get("sheetPageSetup") or []

    if not template_name:
        return jsonify({"error": "templateName is required"}), 400
    if output_format not in {"excel", "pdf"}:
        return jsonify({"error": "outputFormat must be excel or pdf"}), 400
    if postprocess not in SUPPORTED_POSTPROCESS:
        return jsonify({"error": "Unsupported postprocess"}), 400
    if not isinstance(cells, list):
        return jsonify({"error": "cells must be an array"}), 400
    if not isinstance(row_visibility, list):
        return jsonify({"error": "rowVisibility must be an array"}), 400
    if not isinstance(row_heights, list):
        return jsonify({"error": "rowHeights must be an array"}), 400
    if not isinstance(row_breaks, list):
        return jsonify({"error": "rowBreaks must be an array"}), 400
    if not isinstance(print_areas, list):
        return jsonify({"error": "printAreas must be an array"}), 400
    if not isinstance(range_copies, list):
        return jsonify({"error": "rangeCopies must be an array"}), 400
    if not isinstance(sheet_copies, list):
        return jsonify({"error": "sheetCopies must be an array"}), 400
    if not isinstance(hidden_sheets, list):
        return jsonify({"error": "hiddenSheets must be an array"}), 400
    if not isinstance(sheet_page_setup, list):
        return jsonify({"error": "sheetPageSetup must be an array"}), 400

    clear_cache_dir()
    temp_dir = Path(tempfile.mkdtemp(prefix="segmentica-render-"))
    try:
        template_path = ensure_template_path(template_name)
        generated_xlsx = temp_dir / f"{file_base_name}.xlsx"
        apply_cells_to_workbook(
            template_path,
            generated_xlsx,
            cells,
            row_visibility,
            row_heights,
            row_breaks,
            print_areas,
            range_copies,
            sheet_copies,
            hidden_sheets,
            sheet_page_setup,
        )
        final_xlsx_path = generated_xlsx
        if should_normalize_excel_output(template_name, output_format):
            final_xlsx_path = normalize_excel_output(generated_xlsx, temp_dir)
        generated_xlsx_bytes = final_xlsx_path.read_bytes()

        if output_format == "excel":
            response = send_file(
                io.BytesIO(generated_xlsx_bytes),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="document.xlsx",
            )
            return apply_download_filename(response, f"{file_base_name}.xlsx", as_attachment=True)

        ensure_cache_dir()
        cache_key = build_cache_key(generated_xlsx_bytes, output_format, postprocess)
        cached_pdf_path = CACHE_DIR / f"{cache_key}.pdf"
        if cached_pdf_path.exists():
            pdf_bytes = cached_pdf_path.read_bytes()
            response = send_file(
                io.BytesIO(pdf_bytes),
                mimetype="application/pdf",
                as_attachment=True,
                download_name="document.pdf",
            )
            return apply_download_filename(response, f"{file_base_name}.pdf", as_attachment=True)

        pdf_path = convert_to_pdf(generated_xlsx, temp_dir)
        final_pdf = pdf_path

        pdf_bytes = final_pdf.read_bytes()
        cached_pdf_path.write_bytes(pdf_bytes)
        response = send_file(
            io.BytesIO(pdf_bytes),
            mimetype="application/pdf",
            as_attachment=True,
            download_name="document.pdf",
        )
        return apply_download_filename(response, f"{file_base_name}.pdf", as_attachment=True)
    except FileNotFoundError as error:
        return jsonify({"error": str(error)}), 404
    except Exception as error:
        traceback.print_exc()
        return jsonify({"error": str(error)}), 500
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


@app.post("/render/docx-template")
def render_docx_template():
    payload = request.get_json(silent=True) or {}
    template_name = str(payload.get("templateName") or "").strip()
    file_base_name = str(payload.get("fileBaseName") or "document").strip() or "document"
    output_format = str(payload.get("outputFormat") or "pdf").strip().lower()
    replacements = payload.get("replacements") or {}
    replace_first_image_base64 = str(payload.get("replaceFirstImageBase64") or "").strip()

    if not template_name:
        return jsonify({"error": "templateName is required"}), 400
    if output_format not in {"word", "pdf"}:
        return jsonify({"error": "outputFormat must be word or pdf"}), 400
    if not isinstance(replacements, dict):
        return jsonify({"error": "replacements must be an object"}), 400

    replace_first_image_bytes = None
    if replace_first_image_base64:
        try:
            replace_first_image_bytes = base64.b64decode(replace_first_image_base64, validate=True)
        except Exception:
            return jsonify({"error": "replaceFirstImageBase64 must be valid base64"}), 400

    clear_cache_dir()
    temp_dir = Path(tempfile.mkdtemp(prefix="segmentica-render-"))
    try:
        template_path = ensure_template_path(template_name)
        normalized_template_path = convert_word_template_to_docx(template_path, temp_dir)
        generated_docx = temp_dir / f"{file_base_name}.docx"
        normalized_replacements = {str(key): "" if value is None else str(value) for key, value in replacements.items()}
        if template_name == "Specifikacia_k_dogovoru_postavki.docx":
            apply_order_supply_specification_template(
                normalized_template_path,
                generated_docx,
                normalized_replacements,
            )
        elif template_name == "Счет_образец.docx":
            apply_structured_invoice_template(
                normalized_template_path,
                generated_docx,
                normalized_replacements,
                replace_first_image_bytes,
            )
        else:
            apply_replacements_to_docx(
                normalized_template_path,
                generated_docx,
                normalized_replacements,
                replace_first_image_bytes,
            )
        generated_docx_bytes = generated_docx.read_bytes()

        if output_format == "word":
            response = send_file(
                io.BytesIO(generated_docx_bytes),
                mimetype=DOCX_MIMETYPE,
                as_attachment=True,
                download_name="document.docx",
            )
            return apply_download_filename(response, f"{file_base_name}.docx", as_attachment=True)

        ensure_cache_dir()
        cache_key = build_cache_key(generated_docx_bytes, output_format, "none")
        cached_pdf_path = CACHE_DIR / f"{cache_key}.pdf"
        if cached_pdf_path.exists():
            pdf_bytes = cached_pdf_path.read_bytes()
            response = send_file(
                io.BytesIO(pdf_bytes),
                mimetype="application/pdf",
                as_attachment=True,
                download_name="document.pdf",
            )
            return apply_download_filename(response, f"{file_base_name}.pdf", as_attachment=True)

        pdf_path = convert_to_pdf(generated_docx, temp_dir)
        pdf_bytes = pdf_path.read_bytes()
        cached_pdf_path.write_bytes(pdf_bytes)
        response = send_file(
            io.BytesIO(pdf_bytes),
            mimetype="application/pdf",
            as_attachment=True,
            download_name="document.pdf",
        )
        return apply_download_filename(response, f"{file_base_name}.pdf", as_attachment=True)
    except FileNotFoundError as error:
        return jsonify({"error": str(error)}), 404
    except Exception as error:
        traceback.print_exc()
        return jsonify({"error": str(error)}), 500
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "3000"))
    app.run(host="0.0.0.0", port=port)
